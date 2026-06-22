# 교수학습과정안 DB(단원별 핵심 내용 마스터 엑셀) → 단원 레코드 추출.
# 국어·수학·사회·과학만. 결과: scripts/units.json (전체 레코드 + 카탈로그 트리)
import openpyxl, os, json, re

BASE = "/Users/jawoon/Library/CloudStorage/OneDrive-서울창도초등학교/@ 2026 질문이 있는 교수학습 과정안 도우미 개발/파일 정리"
FILES = [
    "(34학년)단원별 핵심 내용 추출_260613_백업_260615.xlsx",
    "(56학년)단원별 핵심 내용 추출_260612_백업_260615.xlsx",
]
SUBJECTS = ["국어", "수학", "사회", "과학"]
PUB_NORM = {"천재(정용재)": "천재교과서(정용재)"}  # 과학 중복표기 통일

records = []
seen = set()
for fn in FILES:
    wb = openpyxl.load_workbook(os.path.join(BASE, fn), read_only=True, data_only=True)
    for sub in SUBJECTS:
        if sub not in wb.sheetnames:
            continue
        ws = wb[sub]
        rows = ws.iter_rows(min_row=2, values_only=True)
        for r in rows:
            if not r or all(c is None for c in r):
                continue
            과목, 학년, 학기, 출판사, 단원명, 성취기준, 단원학습내용, 영역 = (list(r) + [None] * 8)[:8]
            if not (과목 and 학년 and 단원명):
                continue
            rec = {
                "subject": str(과목).strip(),
                "grade": str(학년).strip(),
                "semester": str(학기).strip() if 학기 is not None else "",
                "publisher": PUB_NORM.get(str(출판사).strip(), str(출판사).strip()) if 출판사 else "",
                "unit": re.sub(r"^\s*\d+\.\s*", "", str(단원명).strip()),  # "1. 곱셈" → "곱셈"
                "unitRaw": str(단원명).strip(),
                "standard": str(성취기준).strip() if 성취기준 else "",
                "content": str(단원학습내용).strip() if 단원학습내용 else "",
                "domain": str(영역).strip() if 영역 else "",
            }
            key = "|".join([rec["subject"], rec["grade"], rec["semester"], rec["publisher"], rec["unitRaw"]])
            if key in seen:
                continue
            seen.add(key)
            rec["key"] = key
            records.append(rec)
    wb.close()

# 카탈로그 트리: 학년 > 학기 > 과목 > 출판사 > [단원]
catalog = {}
for r in records:
    g = catalog.setdefault(r["grade"], {})
    s = g.setdefault(r["semester"], {})
    sub = s.setdefault(r["subject"], {})
    pub = sub.setdefault(r["publisher"], [])
    pub.append({"unit": r["unit"], "unitRaw": r["unitRaw"], "key": r["key"]})

out = {"records": records, "catalog": catalog}
with open(os.path.join(os.path.dirname(__file__), "units.json"), "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=2)

# 요약
print(f"총 단원 레코드: {len(records)}")
from collections import Counter
bysub = Counter(r["subject"] for r in records)
print("과목별:", dict(bysub))
for sub in SUBJECTS:
    pubs = sorted(set(r["publisher"] for r in records if r["subject"] == sub and r["publisher"]))
    print(f"  [{sub}] 출판사 {len(pubs)}종: {pubs}")
empty = sum(1 for r in records if not r["content"])
print(f"단원학습내용 비어있는 레코드: {empty}")
print("저장: scripts/units.json")
